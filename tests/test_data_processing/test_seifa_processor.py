"""
Comprehensive unit tests for SEIFA (Socio-Economic Indexes for Areas) processor.

Tests the real Excel data processing pipeline with mock data generation,
validation, error handling, and edge case scenarios.

Covers:
- Excel file validation and structure checking
- Data extraction from Table 1 sheet
- Column standardization and type conversion
- SA2 code validation (9-digit Australian codes)
- SEIFA index validation (scores 800-1200, deciles 1-10)
- Error handling for corrupted/invalid data
- Performance benchmarks for large datasets
"""

import pytest
import polars as pl
import numpy as np
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
import tempfile
import shutil
from openpyxl import Workbook
import time

from src.data_processing.seifa_processor import (
    SEIFAProcessor, 
    SEIFA_CONFIG, 
    SEIFA_SCHEMA, 
    SEIFA_INDEX_DESCRIPTIONS
)


class TestSEIFAProcessor:
    """Comprehensive test suite for SEIFA processor."""
    
    def test_seifa_processor_initialization(self, mock_data_paths):
        """Test SEIFA processor initializes correctly with proper directory structure."""
        processor = SEIFAProcessor(data_dir=mock_data_paths["raw_dir"].parent)
        
        assert processor.data_dir == mock_data_paths["raw_dir"].parent
        assert processor.raw_dir.exists()
        assert processor.processed_dir.exists()
        assert processor.raw_dir.name == "raw"
        assert processor.processed_dir.name == "processed"
    
    def test_seifa_processor_default_directory(self):
        """Test SEIFA processor with default data directory."""
        with tempfile.TemporaryDirectory() as temp_dir:
            original_cwd = Path.cwd()
            try:
                # Change to temp directory
                import os
                os.chdir(temp_dir)
                
                processor = SEIFAProcessor()
                assert processor.data_dir.name == "data"
                assert processor.raw_dir.exists()
                assert processor.processed_dir.exists()
            finally:
                import os
                os.chdir(original_cwd)
    
    def test_validate_seifa_file_missing(self, mock_seifa_processor):
        """Test validation fails for missing SEIFA file."""
        non_existent_file = Path("non_existent_file.xlsx")
        result = mock_seifa_processor.validate_seifa_file(non_existent_file)
        assert result is False
    
    def test_validate_seifa_file_valid(self, mock_excel_seifa_file, mock_seifa_processor):
        """Test validation passes for properly structured SEIFA file."""
        excel_file = mock_excel_seifa_file(num_areas=50)
        result = mock_seifa_processor.validate_seifa_file(excel_file)
        assert result is True
    
    def test_validate_seifa_file_missing_sheet(self, mock_data_paths, mock_seifa_processor):
        """Test validation fails when primary sheet is missing."""
        # Create Excel file without Table 1 sheet
        filepath = mock_data_paths["seifa_excel"]
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Contents")
        wb.create_sheet("Wrong Sheet")
        wb.save(filepath)
        
        result = mock_seifa_processor.validate_seifa_file(filepath)
        assert result is False
    
    def test_validate_seifa_file_size_warning(self, mock_data_paths, mock_seifa_processor, caplog):
        """Test file size validation generates appropriate warnings."""
        # Create very small file
        filepath = mock_data_paths["seifa_excel"]
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Table 1")
        wb.save(filepath)
        
        with caplog.at_level("WARNING"):
            mock_seifa_processor.validate_seifa_file(filepath)
            assert "Unexpected SEIFA file size" in caplog.text
    
    def test_extract_seifa_data_valid_file(self, mock_excel_seifa_file, mock_seifa_processor):
        """Test successful extraction from valid SEIFA Excel file."""
        excel_file = mock_excel_seifa_file(num_areas=20)
        
        result_df = mock_seifa_processor.extract_seifa_data(excel_file)
        
        # Verify basic structure
        assert isinstance(result_df, pl.DataFrame)
        assert len(result_df) > 0
        assert len(result_df) <= 20  # Should match or be less due to validation
        
        # Verify required columns exist
        expected_columns = ["sa2_code_2021", "sa2_name_2021"]
        for col in expected_columns:
            assert col in result_df.columns
    
    def test_extract_seifa_data_with_errors(self, mock_excel_seifa_file, mock_seifa_processor):
        """Test extraction handles Excel files with data errors gracefully."""
        excel_file = mock_excel_seifa_file(num_areas=30, include_errors=True)
        
        result_df = mock_seifa_processor.extract_seifa_data(excel_file)
        
        # Should still extract data despite errors
        assert isinstance(result_df, pl.DataFrame)
        assert len(result_df) > 0
        
        # Verify error handling preserved data integrity
        if "sa2_code_2021" in result_df.columns:
            codes = result_df["sa2_code_2021"].to_list()
            valid_codes = [code for code in codes if code and len(str(code)) == 9]
            assert len(valid_codes) > 0
    
    def test_standardize_seifa_columns(self, mock_seifa_processor, mock_seifa_data):
        """Test SEIFA column standardization process."""
        # Create test data with Excel-like structure
        original_df = mock_seifa_data(num_areas=10)
        
        # Rename columns to simulate Excel extraction
        excel_columns = [
            "SA2 Code", "SA2 Name", "IRSD Score", "IRSD Decile",
            "IRSAD Score", "IRSAD Decile", "IER Score", "IER Decile", 
            "IEO Score", "IEO Decile", "Population"
        ]
        
        # Create DataFrame with Excel-like column names
        excel_df = original_df.select(original_df.columns[:11])
        excel_df.columns = excel_columns
        
        standardized_df = mock_seifa_processor._standardize_seifa_columns(excel_df)
        
        # Verify standardized column names
        expected_columns = [
            "sa2_code_2021", "sa2_name_2021", "irsd_score", "irsd_decile",
            "irsad_score", "irsad_decile", "ier_score", "ier_decile",
            "ieo_score", "ieo_decile", "usual_resident_population"
        ]
        
        for col in expected_columns:
            if col in SEIFA_SCHEMA:
                assert col in standardized_df.columns
    
    def test_standardize_seifa_columns_insufficient_columns(self, mock_seifa_processor):
        """Test standardization fails gracefully with insufficient columns."""
        # Create DataFrame with too few columns
        df = pl.DataFrame({"col1": [1, 2, 3], "col2": ["a", "b", "c"]})
        
        with pytest.raises(ValueError, match="Expected at least 11 columns"):
            mock_seifa_processor._standardize_seifa_columns(df)
    
    def test_validate_seifa_data_valid(self, mock_seifa_processor, mock_seifa_data):
        """Test SEIFA data validation with valid data."""
        valid_df = mock_seifa_data(num_areas=15)
        
        validated_df = mock_seifa_processor._validate_seifa_data(valid_df)
        
        assert isinstance(validated_df, pl.DataFrame)
        assert len(validated_df) > 0
        
        # Verify SA2 codes are valid
        if "sa2_code_2021" in validated_df.columns:
            codes = validated_df["sa2_code_2021"].to_list()
            for code in codes:
                if code:  # Skip nulls
                    assert len(str(code)) == 9
                    assert str(code).isdigit()
    
    def test_validate_seifa_data_invalid_sa2_codes(self, mock_seifa_processor, error_simulation):
        """Test validation filters out invalid SA2 codes."""
        # Create DataFrame with invalid SA2 codes
        df = pl.DataFrame({
            "sa2_code_2021": ["123456789", "12345", "INVALID", "987654321"],
            "sa2_name_2021": ["Valid Area 1", "Invalid Area 1", "Invalid Area 2", "Valid Area 2"],
            "irsd_score": [900, 950, 1000, 1050],
            "irsd_decile": [3, 4, 5, 6]
        })
        
        validated_df = mock_seifa_processor._validate_seifa_data(df)
        
        # Should only keep valid 9-digit codes
        assert len(validated_df) == 2
        codes = validated_df["sa2_code_2021"].to_list()
        for code in codes:
            assert len(code) == 9
            assert code.isdigit()
    
    def test_validate_seifa_data_invalid_scores(self, mock_seifa_processor):
        """Test validation filters out invalid SEIFA scores."""
        df = pl.DataFrame({
            "sa2_code_2021": ["123456789", "234567890", "345678901", "456789012"],
            "sa2_name_2021": ["Area 1", "Area 2", "Area 3", "Area 4"],
            "irsd_score": [900, 500, 1500, 1000],  # 500 and 1500 are invalid
            "irsd_decile": [3, 4, 5, 6]
        })
        
        validated_df = mock_seifa_processor._validate_seifa_data(df)
        
        # Should filter out rows with invalid scores
        assert len(validated_df) == 2
        scores = validated_df["irsd_score"].to_list()
        for score in scores:
            if score is not None:
                assert 800 <= score <= 1200
    
    def test_validate_seifa_data_invalid_deciles(self, mock_seifa_processor):
        """Test validation filters out invalid SEIFA deciles."""
        df = pl.DataFrame({
            "sa2_code_2021": ["123456789", "234567890", "345678901", "456789012"],
            "sa2_name_2021": ["Area 1", "Area 2", "Area 3", "Area 4"],
            "irsd_score": [900, 950, 1000, 1050],
            "irsd_decile": [3, 0, 15, 6]  # 0 and 15 are invalid
        })
        
        validated_df = mock_seifa_processor._validate_seifa_data(df)
        
        # Should filter out rows with invalid deciles
        assert len(validated_df) == 2
        deciles = validated_df["irsd_decile"].to_list()
        for decile in deciles:
            if decile is not None:
                assert 1 <= decile <= 10
    
    def test_validate_seifa_data_with_nulls(self, mock_seifa_processor):
        """Test validation handles null values appropriately."""
        df = pl.DataFrame({
            "sa2_code_2021": ["123456789", "234567890", "345678901"],
            "sa2_name_2021": ["Area 1", "Area 2", "Area 3"], 
            "irsd_score": [900, None, 1000],
            "irsd_decile": [3, None, 6],
            "irsad_score": [None, 950, 1050]
        })
        
        validated_df = mock_seifa_processor._validate_seifa_data(df)
        
        # Should keep all rows since nulls are allowed
        assert len(validated_df) == 3
        
        # Verify null handling
        assert validated_df["irsd_score"].null_count() == 1
        assert validated_df["irsad_score"].null_count() == 1
    
    def test_validate_seifa_data_significant_loss_warning(self, mock_seifa_processor, caplog):
        """Test warning is generated when significant data loss occurs during validation."""
        # Create mostly invalid data
        df = pl.DataFrame({
            "sa2_code_2021": ["INVALID"] * 2000 + ["123456789"] * 100,  # Only 100 valid out of 2100
            "sa2_name_2021": [f"Area {i}" for i in range(2100)],
            "irsd_score": [900] * 2100,
            "irsd_decile": [5] * 2100
        })
        
        with caplog.at_level("WARNING"):
            validated_df = mock_seifa_processor._validate_seifa_data(df)
            assert "Significant data loss during validation" in caplog.text
        
        # Should only have 100 valid records
        assert len(validated_df) == 100
    
    def test_process_seifa_file_complete_pipeline(self, mock_excel_seifa_file, mock_seifa_processor):
        """Test complete SEIFA file processing pipeline."""
        excel_file = mock_excel_seifa_file(num_areas=25)
        
        # Copy file to expected location
        expected_path = mock_seifa_processor.raw_dir / SEIFA_CONFIG["filename"]
        shutil.copy(excel_file, expected_path)
        
        result_df = mock_seifa_processor.process_seifa_file()
        
        # Verify processing succeeded
        assert isinstance(result_df, pl.DataFrame)
        assert len(result_df) > 0
        
        # Verify output files were created
        csv_output = mock_seifa_processor.processed_dir / "seifa_2021_sa2.csv"
        parquet_output = mock_seifa_processor.processed_dir / "seifa_2021_sa2.parquet"
        
        assert csv_output.exists()
        assert parquet_output.exists()
        
        # Verify output file contents
        csv_df = pl.read_csv(csv_output)
        parquet_df = pl.read_parquet(parquet_output)
        
        assert len(csv_df) == len(result_df)
        assert len(parquet_df) == len(result_df)
    
    def test_process_seifa_file_missing_file(self, mock_seifa_processor):
        """Test processing fails gracefully when SEIFA file is missing."""
        with pytest.raises(FileNotFoundError):
            mock_seifa_processor.process_seifa_file()
    
    def test_process_seifa_file_custom_filename(self, mock_excel_seifa_file, mock_seifa_processor):
        """Test processing with custom filename."""
        custom_filename = "custom_seifa.xlsx"
        excel_file = mock_excel_seifa_file(num_areas=10)
        
        # Copy to custom location
        custom_path = mock_seifa_processor.raw_dir / custom_filename
        shutil.copy(excel_file, custom_path)
        
        result_df = mock_seifa_processor.process_seifa_file(filename=custom_filename)
        
        assert isinstance(result_df, pl.DataFrame)
        assert len(result_df) > 0
    
    def test_get_seifa_summary_comprehensive(self, mock_seifa_processor, mock_seifa_data):
        """Test comprehensive SEIFA data summary generation."""
        seifa_df = mock_seifa_data(num_areas=50)
        
        summary = mock_seifa_processor.get_seifa_summary(seifa_df)
        
        # Verify summary structure
        assert "total_sa2_areas" in summary
        assert "states_covered" in summary
        assert "seifa_statistics" in summary
        
        # Verify counts
        assert summary["total_sa2_areas"] == len(seifa_df)
        
        # Verify state coverage
        assert isinstance(summary["states_covered"], list)
        assert all(state in ["NSW", "VIC", "QLD", "SA", "WA", "TAS", "NT", "ACT", "Unknown(9)"] 
                  for state in summary["states_covered"])
        
        # Verify SEIFA statistics
        stats = summary["seifa_statistics"]
        expected_indices = ["IRSD", "IRSAD", "IER", "IEO"]
        
        for index in expected_indices:
            if index in stats:
                index_stats = stats[index]
                assert "min_score" in index_stats
                assert "max_score" in index_stats
                assert "mean_score" in index_stats
                assert "median_score" in index_stats
                
                # Verify reasonable ranges
                if index_stats["min_score"] is not None:
                    assert 800 <= index_stats["min_score"] <= 1200
                if index_stats["max_score"] is not None:
                    assert 800 <= index_stats["max_score"] <= 1200
    
    def test_get_seifa_summary_missing_columns(self, mock_seifa_processor):
        """Test summary generation with missing columns."""
        # DataFrame with minimal columns
        df = pl.DataFrame({
            "sa2_code_2021": ["123456789", "234567890"],
            "sa2_name_2021": ["Area 1", "Area 2"]
        })
        
        summary = mock_seifa_processor.get_seifa_summary(df)
        
        assert summary["total_sa2_areas"] == 2
        assert "states_covered" in summary
        assert "seifa_statistics" in summary
    
    def test_process_complete_pipeline_success(self, mock_excel_seifa_file, mock_seifa_processor):
        """Test complete processing pipeline executes successfully."""
        excel_file = mock_excel_seifa_file(num_areas=30)
        
        # Copy file to expected location
        expected_path = mock_seifa_processor.raw_dir / SEIFA_CONFIG["filename"]
        shutil.copy(excel_file, expected_path)
        
        result_df = mock_seifa_processor.process_complete_pipeline()
        
        assert isinstance(result_df, pl.DataFrame)
        assert len(result_df) > 0
        
        # Verify processing outputs
        assert (mock_seifa_processor.processed_dir / "seifa_2021_sa2.csv").exists()
        assert (mock_seifa_processor.processed_dir / "seifa_2021_sa2.parquet").exists()
    
    def test_process_complete_pipeline_failure(self, mock_seifa_processor):
        """Test complete pipeline handles failures gracefully."""
        # No SEIFA file exists
        with pytest.raises(FileNotFoundError):
            mock_seifa_processor.process_complete_pipeline()
    
    def test_performance_large_dataset(self, mock_excel_seifa_file, mock_seifa_processor):
        """Test processing performance with large dataset."""
        # Create larger dataset for performance testing
        excel_file = mock_excel_seifa_file(num_areas=1000)
        
        expected_path = mock_seifa_processor.raw_dir / SEIFA_CONFIG["filename"]
        shutil.copy(excel_file, expected_path)
        
        start_time = time.time()
        result_df = mock_seifa_processor.process_seifa_file()
        processing_time = time.time() - start_time
        
        # Performance assertions
        assert processing_time < 10.0  # Should complete within 10 seconds
        assert isinstance(result_df, pl.DataFrame)
        assert len(result_df) > 0
        
        # Memory efficiency check
        memory_usage = result_df.estimated_size("mb")
        assert memory_usage < 100  # Should not exceed 100MB for 1000 records
    
    def test_data_type_consistency(self, mock_seifa_processor, mock_seifa_data):
        """Test data type consistency throughout processing pipeline."""
        seifa_df = mock_seifa_data(num_areas=20)
        
        # Process through standardization
        standardized_df = mock_seifa_processor._standardize_seifa_columns(seifa_df)
        
        # Verify expected data types
        for col, expected_type in SEIFA_SCHEMA.items():
            if col in standardized_df.columns:
                actual_type = standardized_df[col].dtype
                # Allow for some type flexibility (e.g., Int32 vs Int64)
                assert actual_type in [expected_type, pl.Int64] or str(actual_type) == str(expected_type)
    
    def test_error_handling_corrupt_excel(self, mock_data_paths, mock_seifa_processor):
        """Test error handling with corrupted Excel file."""
        # Create invalid Excel file
        corrupt_file = mock_data_paths["seifa_excel"]
        corrupt_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Write invalid content
        with open(corrupt_file, "w") as f:
            f.write("This is not an Excel file")
        
        with pytest.raises(Exception):
            mock_seifa_processor.extract_seifa_data(corrupt_file)
    
    def test_edge_case_empty_excel_sheet(self, mock_data_paths, mock_seifa_processor):
        """Test handling of empty Excel sheet."""
        filepath = mock_data_paths["seifa_excel"]
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # Create Excel with empty Table 1 sheet
        wb = Workbook()
        wb.remove(wb.active)
        table1 = wb.create_sheet("Table 1")
        
        # Add headers but no data
        headers = ["SA2 Code", "SA2 Name", "IRSD Score", "IRSD Decile"]
        for col, header in enumerate(headers, 1):
            table1.cell(row=6, column=col, value=header)
        
        wb.save(filepath)
        
        result_df = mock_seifa_processor.extract_seifa_data(filepath)
        
        # Should handle empty data gracefully
        assert isinstance(result_df, pl.DataFrame)
        assert len(result_df) == 0  # No data rows
    
    def test_concurrent_processing_safety(self, mock_excel_seifa_file, mock_data_paths):
        """Test thread safety of SEIFA processor (basic check)."""
        import threading
        import concurrent.futures
        
        excel_file = mock_excel_seifa_file(num_areas=10)
        
        def process_seifa():
            processor = SEIFAProcessor(data_dir=mock_data_paths["raw_dir"].parent)
            expected_path = processor.raw_dir / SEIFA_CONFIG["filename"]
            shutil.copy(excel_file, expected_path)
            return processor.process_seifa_file()
        
        # Run multiple processors concurrently
        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
            futures = [executor.submit(process_seifa) for _ in range(3)]
            results = [future.result() for future in concurrent.futures.as_completed(futures)]
        
        # All should succeed
        assert len(results) == 3
        for result in results:
            assert isinstance(result, pl.DataFrame)
            assert len(result) > 0


class TestSEIFAConfiguration:
    """Test SEIFA configuration constants and schema."""
    
    def test_seifa_config_structure(self):
        """Test SEIFA configuration has expected structure."""
        required_keys = [
            "filename", "primary_sheet", "header_row", "data_start_row",
            "expected_records", "expected_sheets"
        ]
        
        for key in required_keys:
            assert key in SEIFA_CONFIG
        
        # Verify data types
        assert isinstance(SEIFA_CONFIG["filename"], str)
        assert isinstance(SEIFA_CONFIG["primary_sheet"], str)
        assert isinstance(SEIFA_CONFIG["header_row"], int)
        assert isinstance(SEIFA_CONFIG["data_start_row"], int)
        assert isinstance(SEIFA_CONFIG["expected_records"], int)
        assert isinstance(SEIFA_CONFIG["expected_sheets"], list)
    
    def test_seifa_schema_completeness(self):
        """Test SEIFA schema covers all expected columns."""
        required_columns = [
            "sa2_code_2021", "sa2_name_2021", "usual_resident_population"
        ]
        
        seifa_indices = ["irsd", "irsad", "ier", "ieo"]
        for index in seifa_indices:
            required_columns.extend([f"{index}_score", f"{index}_decile"])
        
        for col in required_columns:
            assert col in SEIFA_SCHEMA
        
        # Verify Polars data types
        for col, dtype in SEIFA_SCHEMA.items():
            assert hasattr(pl, dtype.__name__) or isinstance(dtype, type)
    
    def test_seifa_index_descriptions(self):
        """Test SEIFA index descriptions are complete."""
        expected_indices = ["irsd", "irsad", "ier", "ieo"]
        
        for index in expected_indices:
            assert index in SEIFA_INDEX_DESCRIPTIONS
            assert isinstance(SEIFA_INDEX_DESCRIPTIONS[index], str)
            assert len(SEIFA_INDEX_DESCRIPTIONS[index]) > 10  # Meaningful description


# ============================================================================
# PROPERTY-BASED TESTING (if hypothesis available)
# ============================================================================

try:
    from hypothesis import given, strategies as st, assume
    import hypothesis.extra.numpy as npst
    
    class TestSEIFAPropertyBased:
        """Property-based tests for SEIFA processor using hypothesis."""
        
        @given(
            sa2_codes=st.lists(
                st.text(alphabet="12345678", min_size=9, max_size=9)
                .filter(lambda x: x[0] in "12345678"),
                min_size=1,
                max_size=50
            ),
            scores=st.lists(
                st.integers(min_value=800, max_value=1200),
                min_size=1,
                max_size=50
            )
        )
        def test_seifa_validation_properties(self, sa2_codes, scores, mock_seifa_processor):
            """Property-based test for SEIFA data validation."""
            assume(len(sa2_codes) == len(scores))
            
            # Create test DataFrame
            df = pl.DataFrame({
                "sa2_code_2021": sa2_codes,
                "sa2_name_2021": [f"Area {i}" for i in range(len(sa2_codes))],
                "irsd_score": scores,
                "irsd_decile": [min(10, max(1, score // 40)) for score in scores]
            })
            
            validated_df = mock_seifa_processor._validate_seifa_data(df)
            
            # Properties that should always hold
            assert len(validated_df) <= len(df)  # Should not gain rows
            
            # All remaining SA2 codes should be valid
            if len(validated_df) > 0 and "sa2_code_2021" in validated_df.columns:
                codes = validated_df["sa2_code_2021"].to_list()
                for code in codes:
                    if code:
                        assert len(code) == 9
                        assert code.isdigit()
                        assert code[0] in "12345678"
            
            # All remaining scores should be in valid range
            if len(validated_df) > 0 and "irsd_score" in validated_df.columns:
                valid_scores = validated_df["irsd_score"].drop_nulls().to_list()
                for score in valid_scores:
                    assert 800 <= score <= 1200

except ImportError:
    # Skip property-based tests if hypothesis not available
    pass