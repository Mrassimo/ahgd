"""
Comprehensive pytest configuration and fixtures for Australian Health Analytics.

Provides mock data generators for Australian health data patterns:
- SA2 codes (9-digit Australian statistical areas)
- SEIFA socio-economic indices (1-10 deciles, 800-1200 scores)
- PBS prescription data
- Geographic boundary data
- Health risk assessment data

Supports property-based testing and performance benchmarking.
"""

import pytest
import polars as pl
import numpy as np
import tempfile
import shutil
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime, date, timedelta
import random
from unittest.mock import Mock, MagicMock
import io
import openpyxl
from openpyxl import Workbook

# Set random seeds for reproducible testing
np.random.seed(42)
random.seed(42)


# ============================================================================
# CONFIGURATION FIXTURES
# ============================================================================

@pytest.fixture(scope="session")
def australian_health_config():
    """Configuration constants for Australian health data patterns."""
    return {
        # Australian SA2 (Statistical Area Level 2) patterns
        "sa2_patterns": {
            "total_sa2_areas": 2454,  # Total SA2 areas in Australia
            "code_length": 9,  # SA2 codes are 9 digits
            "state_prefixes": {
                "1": "NSW",  # New South Wales
                "2": "VIC",  # Victoria
                "3": "QLD",  # Queensland
                "4": "SA",   # South Australia
                "5": "WA",   # Western Australia
                "6": "TAS",  # Tasmania
                "7": "NT",   # Northern Territory
                "8": "ACT"   # Australian Capital Territory
            }
        },
        
        # SEIFA (Socio-Economic Indexes for Areas) patterns
        "seifa_patterns": {
            "decile_range": (1, 10),      # SEIFA deciles 1-10
            "score_range": (800, 1200),   # SEIFA scores typically 800-1200
            "expected_records": 2368,     # Expected SA2 areas with SEIFA data
            "indices": ["irsd", "irsad", "ier", "ieo"]  # Four SEIFA indices
        },
        
        # PBS (Pharmaceutical Benefits Scheme) patterns
        "pbs_patterns": {
            "atc_codes": [
                "A02BC01", "A10BD07", "C07AB02", "C09AA02", "J01CA04",
                "N02BE01", "R03AC02", "C08CA05", "A03FA01", "H03AA01"
            ],  # Common ATC codes
            "prescription_range": (1, 50),    # Prescriptions per SA2
            "cost_range": (10.0, 500.0),     # Cost range AUD
            "chronic_medication_rate": 0.3    # 30% chronic medications
        },
        
        # Geographic patterns
        "geographic_patterns": {
            "area_range": (0.1, 1000.0),     # SA2 area in km²
            "population_range": (100, 25000), # Population per SA2
            "remoteness_categories": [
                "Major Cities", "Inner Regional", "Outer Regional", 
                "Remote", "Very Remote"
            ]
        },
        
        # Health risk patterns
        "risk_patterns": {
            "risk_categories": ["Very Low", "Low", "Medium", "High", "Very High"],
            "access_categories": ["Excellent", "Good", "Fair", "Poor", "Very Poor"],
            "composite_score_range": (0.0, 100.0),
            "seifa_risk_weight": 0.6,
            "health_util_weight": 0.4
        }
    }


@pytest.fixture(scope="session")
def temp_data_directory():
    """Create temporary directory for test data files."""
    temp_dir = tempfile.mkdtemp(prefix="health_analytics_test_")
    yield Path(temp_dir)
    shutil.rmtree(temp_dir)


@pytest.fixture
def mock_data_paths(temp_data_directory):
    """Create mock file paths for testing."""
    paths = {
        "raw_dir": temp_data_directory / "raw",
        "processed_dir": temp_data_directory / "processed",
        "parquet_dir": temp_data_directory / "parquet",
        "seifa_excel": temp_data_directory / "raw" / "SEIFA_2021_SA2_Indexes.xlsx",
        "health_csv": temp_data_directory / "raw" / "PBS_Item_Report_2016_Current.csv",
        "boundaries_zip": temp_data_directory / "raw" / "SA2_2021_boundaries.zip"
    }
    
    # Create directories
    for path_key, path_value in paths.items():
        if path_key.endswith("_dir"):
            path_value.mkdir(parents=True, exist_ok=True)
    
    return paths


# ============================================================================
# MOCK DATA GENERATORS
# ============================================================================

@pytest.fixture
def mock_sa2_codes(australian_health_config):
    """Generate realistic Australian SA2 codes."""
    def generate_codes(count: int = 100) -> List[str]:
        """Generate list of valid 9-digit SA2 codes."""
        codes = []
        state_prefixes = list(australian_health_config["sa2_patterns"]["state_prefixes"].keys())
        
        for _ in range(count):
            state_prefix = np.random.choice(state_prefixes)
            # Generate 8 remaining digits
            remaining_digits = str(np.random.randint(10000000, 99999999))
            sa2_code = f"{state_prefix}{remaining_digits}"
            codes.append(sa2_code)
        
        return codes
    
    return generate_codes


@pytest.fixture
def mock_seifa_data(mock_sa2_codes, australian_health_config):
    """Generate mock SEIFA data with realistic patterns."""
    def generate_seifa(
        num_areas: int = 100,
        with_missing_data: bool = False,
        correlation_strength: float = 0.7
    ) -> pl.DataFrame:
        """
        Generate mock SEIFA data with realistic correlations between indices.
        
        Args:
            num_areas: Number of SA2 areas to generate
            with_missing_data: Include some missing values for testing
            correlation_strength: Correlation between SEIFA indices (0-1)
        """
        config = australian_health_config["seifa_patterns"]
        sa2_codes = mock_sa2_codes(num_areas)
        
        # Generate correlated SEIFA scores
        base_scores = np.random.normal(1000, 100, num_areas)
        base_scores = np.clip(base_scores, config["score_range"][0], config["score_range"][1])
        
        data = {
            "sa2_code_2021": sa2_codes,
            "sa2_name_2021": [f"Mock SA2 Area {i+1}" for i in range(num_areas)]
        }
        
        # Generate correlated indices
        for idx, index_name in enumerate(config["indices"]):
            # Add correlation and noise
            noise = np.random.normal(0, 50, num_areas)
            scores = base_scores + (correlation_strength * noise)
            scores = np.clip(scores, config["score_range"][0], config["score_range"][1])
            
            # Convert to deciles
            deciles = np.ceil((scores - config["score_range"][0]) / 
                            (config["score_range"][1] - config["score_range"][0]) * 10)
            deciles = np.clip(deciles, 1, 10).astype(int)
            
            data[f"{index_name}_score"] = scores.astype(int)
            data[f"{index_name}_decile"] = deciles
        
        # Add population data
        data["usual_resident_population"] = np.random.randint(
            australian_health_config["geographic_patterns"]["population_range"][0],
            australian_health_config["geographic_patterns"]["population_range"][1],
            num_areas
        )
        
        # Introduce missing data if requested
        if with_missing_data:
            missing_rate = 0.05  # 5% missing data
            for col in data:
                if col not in ["sa2_code_2021", "sa2_name_2021"]:
                    mask = np.random.random(num_areas) < missing_rate
                    data[col] = [None if mask[i] else data[col][i] for i in range(num_areas)]
        
        return pl.DataFrame(data)
    
    return generate_seifa


@pytest.fixture
def mock_health_data(mock_sa2_codes, australian_health_config):
    """Generate mock PBS health utilisation data."""
    def generate_health(
        num_records: int = 1000,
        num_sa2_areas: int = 50,
        time_range_days: int = 365
    ) -> pl.DataFrame:
        """
        Generate mock PBS prescription data.
        
        Args:
            num_records: Number of prescription records
            num_sa2_areas: Number of unique SA2 areas
            time_range_days: Date range for prescriptions
        """
        config = australian_health_config["pbs_patterns"]
        sa2_codes = mock_sa2_codes(num_sa2_areas)
        
        # Generate prescription data
        start_date = date.today() - timedelta(days=time_range_days)
        
        data = {
            "sa2_code": np.random.choice(sa2_codes, num_records),
            "atc_code": np.random.choice(config["atc_codes"], num_records),
            "drug_name": [f"Mock Drug {i%20 + 1}" for i in range(num_records)],
            "prescription_count": np.random.randint(
                config["prescription_range"][0], 
                config["prescription_range"][1], 
                num_records
            ),
            "cost_government": np.random.uniform(
                config["cost_range"][0],
                config["cost_range"][1],
                num_records
            ),
            "cost_patient": np.random.uniform(10.0, 50.0, num_records),
            "chronic_medication": np.random.choice(
                [0, 1], 
                num_records, 
                p=[1-config["chronic_medication_rate"], config["chronic_medication_rate"]]
            ),
            "dispensing_date": [
                start_date + timedelta(days=int(np.random.uniform(0, time_range_days)))
                for _ in range(num_records)
            ],
            "state": np.random.choice(
                list(australian_health_config["sa2_patterns"]["state_prefixes"].values()),
                num_records
            )
        }
        
        return pl.DataFrame(data)
    
    return generate_health


@pytest.fixture
def mock_boundary_data(mock_sa2_codes, australian_health_config):
    """Generate mock geographic boundary data."""
    def generate_boundaries(num_areas: int = 100) -> pl.DataFrame:
        """Generate mock SA2 boundary data with geographic attributes."""
        config = australian_health_config["geographic_patterns"]
        sa2_codes = mock_sa2_codes(num_areas)
        
        data = {
            "sa2_code_2021": sa2_codes,
            "sa2_name_2021": [f"Mock SA2 {i+1}" for i in range(num_areas)],
            "state_name": np.random.choice(
                list(australian_health_config["sa2_patterns"]["state_prefixes"].values()),
                num_areas
            ),
            "area_sqkm": np.random.uniform(
                config["area_range"][0],
                config["area_range"][1],
                num_areas
            ),
            "population_2021": np.random.randint(
                config["population_range"][0],
                config["population_range"][1],
                num_areas
            ),
            "remoteness_category": np.random.choice(
                config["remoteness_categories"],
                num_areas
            ),
            # Mock geometry as simple coordinates
            "centroid_lat": np.random.uniform(-44.0, -10.0, num_areas),  # Australian latitude range
            "centroid_lon": np.random.uniform(113.0, 154.0, num_areas),  # Australian longitude range
        }
        
        return pl.DataFrame(data)
    
    return generate_boundaries


@pytest.fixture
def mock_excel_seifa_file(mock_data_paths, mock_seifa_data):
    """Create mock SEIFA Excel file with realistic structure."""
    def create_excel_file(
        filepath: Optional[Path] = None,
        num_areas: int = 100,
        include_errors: bool = False
    ) -> Path:
        """Create a mock SEIFA Excel file matching the real structure."""
        if filepath is None:
            filepath = mock_data_paths["seifa_excel"]
        
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook with expected sheets
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create expected sheets
        sheets = ["Contents", "Table 1", "Table 2", "Table 3", "Table 4", "Table 5", "Table 6"]
        for sheet_name in sheets:
            ws = wb.create_sheet(sheet_name)
            
            # Only populate Table 1 with data
            if sheet_name == "Table 1":
                # Create headers on row 6 (0-indexed: row 5)
                headers = [
                    "SA2 Code", "SA2 Name", "IRSD Score", "IRSD Decile",
                    "IRSAD Score", "IRSAD Decile", "IER Score", "IER Decile",
                    "IEO Score", "IEO Decile", "Population"
                ]
                
                for col, header in enumerate(headers, 1):
                    ws.cell(row=6, column=col, value=header)
                
                # Generate mock data
                seifa_df = mock_seifa_data(num_areas, with_missing_data=include_errors)
                
                # Write data starting from row 7
                for row_idx, row in enumerate(seifa_df.rows(), 7):
                    for col_idx, value in enumerate(row, 1):
                        if include_errors and np.random.random() < 0.02:  # 2% error rate
                            value = "-"  # Simulate missing data
                        ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filepath)
        return filepath
    
    return create_excel_file


# ============================================================================
# COMPONENT FIXTURES
# ============================================================================

@pytest.fixture
def mock_seifa_processor(mock_data_paths):
    """Mock SEIFA processor with temporary data paths."""
    from src.data_processing.seifa_processor import SEIFAProcessor
    return SEIFAProcessor(data_dir=mock_data_paths["raw_dir"].parent)


@pytest.fixture
def mock_health_processor(mock_data_paths):
    """Mock health data processor with temporary data paths."""
    from src.data_processing.health_processor import HealthDataProcessor
    return HealthDataProcessor(data_dir=mock_data_paths["raw_dir"].parent)


@pytest.fixture
def mock_parquet_storage(mock_data_paths):
    """Mock Parquet storage manager with temporary paths."""
    from src.data_processing.storage.parquet_storage_manager import ParquetStorageManager
    return ParquetStorageManager(base_path=mock_data_paths["parquet_dir"])


@pytest.fixture
def mock_risk_calculator(mock_data_paths):
    """Mock health risk calculator with temporary data paths."""
    from src.analysis.risk.health_risk_calculator import HealthRiskCalculator
    return HealthRiskCalculator(data_dir=mock_data_paths["processed_dir"])


# ============================================================================
# PERFORMANCE TESTING FIXTURES
# ============================================================================

@pytest.fixture
def performance_benchmarks():
    """Performance benchmarks for optimization components."""
    return {
        "parquet_compression": {
            "min_compression_ratio": 0.6,  # At least 60% compression
            "max_read_time_per_mb": 0.1,   # Max 0.1s per MB to read
            "max_write_time_per_mb": 0.2   # Max 0.2s per MB to write
        },
        "memory_optimization": {
            "max_memory_increase": 1.5,    # Max 50% memory increase during processing
            "min_memory_reduction": 0.3,   # At least 30% reduction after optimization
        },
        "lazy_loading": {
            "max_initial_load_time": 0.05, # Max 50ms for lazy loader initialization
            "cache_hit_ratio": 0.8,        # At least 80% cache hit ratio
        }
    }


@pytest.fixture
def memory_profiler():
    """Simple memory profiling utility for tests."""
    import psutil
    import os
    
    class MemoryProfiler:
        def __init__(self):
            self.process = psutil.Process(os.getpid())
            self.initial_memory = None
            
        def start(self):
            self.initial_memory = self.process.memory_info().rss / 1024 / 1024  # MB
            
        def get_current_usage(self):
            return self.process.memory_info().rss / 1024 / 1024  # MB
            
        def get_peak_usage(self):
            return self.process.memory_info().peak_wss / 1024 / 1024 if hasattr(
                self.process.memory_info(), 'peak_wss'
            ) else self.get_current_usage()
            
        def get_increase(self):
            if self.initial_memory is None:
                return 0
            return self.get_current_usage() - self.initial_memory
    
    return MemoryProfiler


# ============================================================================
# PROPERTY-BASED TESTING FIXTURES
# ============================================================================

@pytest.fixture
def property_test_strategies():
    """Strategies for property-based testing with hypothesis."""
    try:
        from hypothesis import strategies as st
        
        return {
            # SA2 code strategy
            "sa2_codes": st.text(
                alphabet="12345678",
                min_size=9,
                max_size=9
            ).filter(lambda x: x[0] in "12345678"),
            
            # SEIFA decile strategy
            "seifa_deciles": st.integers(min_value=1, max_value=10),
            
            # SEIFA score strategy
            "seifa_scores": st.integers(min_value=800, max_value=1200),
            
            # Prescription count strategy
            "prescription_counts": st.integers(min_value=0, max_value=100),
            
            # Cost strategy
            "costs": st.floats(min_value=0.0, max_value=1000.0, allow_nan=False),
            
            # Date strategy
            "dates": st.dates(
                min_value=date(2020, 1, 1),
                max_value=date(2024, 12, 31)
            )
        }
    except ImportError:
        # Fallback if hypothesis not available
        return {}


# ============================================================================
# ERROR SIMULATION FIXTURES
# ============================================================================

@pytest.fixture
def error_simulation():
    """Utilities for simulating various error conditions."""
    class ErrorSimulator:
        @staticmethod
        def corrupt_sa2_codes(df: pl.DataFrame, corruption_rate: float = 0.1) -> pl.DataFrame:
            """Introduce invalid SA2 codes for testing error handling."""
            if "sa2_code_2021" not in df.columns and "sa2_code" not in df.columns:
                return df
                
            col_name = "sa2_code_2021" if "sa2_code_2021" in df.columns else "sa2_code"
            corrupted_df = df.clone()
            
            # Randomly corrupt some SA2 codes
            for i in range(len(df)):
                if np.random.random() < corruption_rate:
                    # Create invalid SA2 codes
                    invalid_codes = ["12345", "INVALID", "", None, "123456789A"]
                    corrupted_df = corrupted_df.with_row_index().with_columns(
                        pl.when(pl.col("index") == i)
                        .then(pl.lit(np.random.choice(invalid_codes)))
                        .otherwise(pl.col(col_name))
                        .alias(col_name)
                    ).drop("index")
            
            return corrupted_df
        
        @staticmethod
        def introduce_missing_values(df: pl.DataFrame, missing_rate: float = 0.1) -> pl.DataFrame:
            """Introduce missing values across all columns."""
            corrupted_df = df.clone()
            
            for col in df.columns:
                # Skip critical identifier columns
                if col in ["sa2_code_2021", "sa2_code"]:
                    continue
                    
                mask = np.random.random(len(df)) < missing_rate
                corrupted_df = corrupted_df.with_row_index().with_columns(
                    pl.when(pl.col("index").is_in(np.where(mask)[0]))
                    .then(None)
                    .otherwise(pl.col(col))
                    .alias(col)
                ).drop("index")
            
            return corrupted_df
        
        @staticmethod
        def create_extreme_values(df: pl.DataFrame) -> pl.DataFrame:
            """Create extreme values to test boundary conditions."""
            extreme_df = df.clone()
            
            # Add extreme SEIFA scores
            for col in df.columns:
                if "score" in col:
                    extreme_df = extreme_df.with_columns(
                        pl.when(pl.col(col).is_not_null())
                        .then(pl.lit(np.random.choice([0, 9999])))  # Extreme values
                        .otherwise(pl.col(col))
                        .alias(col)
                    )
                elif "decile" in col:
                    extreme_df = extreme_df.with_columns(
                        pl.when(pl.col(col).is_not_null())
                        .then(pl.lit(np.random.choice([0, 15])))  # Invalid deciles
                        .otherwise(pl.col(col))
                        .alias(col)
                    )
            
            return extreme_df
    
    return ErrorSimulator()


# ============================================================================
# TEST UTILITIES
# ============================================================================

@pytest.fixture
def test_utilities():
    """Common testing utilities and assertion helpers."""
    class TestUtils:
        @staticmethod
        def assert_valid_sa2_codes(codes: List[str]) -> bool:
            """Assert all SA2 codes are valid 9-digit strings."""
            for code in codes:
                if not isinstance(code, str) or len(code) != 9 or not code.isdigit():
                    return False
                if code[0] not in "12345678":  # Valid state prefixes
                    return False
            return True
        
        @staticmethod
        def assert_valid_seifa_deciles(deciles: List[int]) -> bool:
            """Assert all SEIFA deciles are in valid range 1-10."""
            return all(1 <= d <= 10 for d in deciles if d is not None)
        
        @staticmethod
        def assert_valid_seifa_scores(scores: List[int]) -> bool:
            """Assert all SEIFA scores are in reasonable range."""
            return all(800 <= s <= 1200 for s in scores if s is not None)
        
        @staticmethod
        def assert_dataframe_quality(df: pl.DataFrame, expected_cols: List[str]) -> Dict[str, Any]:
            """Comprehensive data frame quality assessment."""
            quality_report = {
                "shape": df.shape,
                "columns_present": all(col in df.columns for col in expected_cols),
                "missing_columns": [col for col in expected_cols if col not in df.columns],
                "null_percentages": {},
                "data_types": dict(zip(df.columns, df.dtypes)),
                "duplicate_rows": df.is_duplicated().sum(),
                "memory_usage_mb": df.estimated_size("mb")
            }
            
            # Calculate null percentages
            for col in df.columns:
                null_count = df.select(pl.col(col).is_null().sum()).item()
                quality_report["null_percentages"][col] = null_count / len(df) * 100
            
            return quality_report
    
    return TestUtils()


# ============================================================================
# INTEGRATION TEST FIXTURES
# ============================================================================

@pytest.fixture
def integration_test_data(mock_seifa_data, mock_health_data, mock_boundary_data):
    """Comprehensive integration test dataset."""
    def create_integration_dataset(
        num_sa2_areas: int = 50,
        num_health_records: int = 500
    ) -> Dict[str, pl.DataFrame]:
        """Create integrated test dataset with consistent SA2 codes."""
        # Generate consistent SA2 codes for all components
        sa2_codes = [f"1{str(i).zfill(8)}" for i in range(10000, 10000 + num_sa2_areas)]
        
        # Create SEIFA data
        seifa_df = mock_seifa_data(num_sa2_areas)
        seifa_df = seifa_df.with_columns(pl.Series("sa2_code_2021", sa2_codes))
        
        # Create health data with same SA2 codes
        health_df = mock_health_data(num_health_records, num_sa2_areas)
        health_df = health_df.with_columns(
            pl.col("sa2_code").map_elements(lambda _: np.random.choice(sa2_codes), return_dtype=pl.Utf8)
        )
        
        # Create boundary data with same SA2 codes
        boundary_df = mock_boundary_data(num_sa2_areas)
        boundary_df = boundary_df.with_columns(pl.Series("sa2_code_2021", sa2_codes))
        
        return {
            "seifa": seifa_df,
            "health": health_df,
            "boundaries": boundary_df,
            "sa2_codes": sa2_codes
        }
    
    return create_integration_dataset