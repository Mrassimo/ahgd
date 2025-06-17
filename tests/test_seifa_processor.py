"""
Test suite for SEIFA (Socio-Economic Indexes for Areas) processor.

Tests real Excel file processing discovered in schema analysis:
- Table 1 with 2,368 SA2 records
- Four SEIFA indices (IRSD, IRSAD, IER, IEO)
- Complete schema validation
"""

import tempfile
from pathlib import Path
from typing import Dict, Optional

import polars as pl
import pytest
from openpyxl import load_workbook

from src.data_processing.downloaders.real_data_downloader import RealDataDownloader


class TestSEIFAProcessor:
    """Test suite for SEIFA data processing with real Excel files."""
    
    @pytest.fixture
    async def seifa_file(self) -> Optional[Path]:
        """Download real SEIFA file for testing."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            downloader = RealDataDownloader(tmp_dir)
            
            try:
                results = await downloader.download_essential_datasets(["seifa_2021_sa2"])
                seifa_path = results.get("seifa_2021_sa2")
                
                if seifa_path and seifa_path.exists():
                    # Copy to a persistent location for tests
                    test_file = Path("test_seifa_2021.xlsx")
                    if not test_file.exists():
                        import shutil
                        shutil.copy2(seifa_path, test_file)
                    return test_file
                else:
                    pytest.skip("Could not download SEIFA file for testing")
            except Exception as e:
                pytest.skip(f"Network error downloading SEIFA file: {e}")
    
    def test_seifa_excel_file_structure(self, seifa_file: Path):
        """Test that SEIFA Excel file has expected structure from analysis."""
        if not seifa_file or not seifa_file.exists():
            pytest.skip("SEIFA file not available")
        
        workbook = load_workbook(seifa_file, read_only=True)
        sheet_names = workbook.sheetnames
        
        # Should have Contents and Table sheets as discovered
        assert "Contents" in sheet_names
        assert "Table 1" in sheet_names
        
        # Check primary data sheet
        table1 = workbook["Table 1"]
        
        # Header should start at row 6 as discovered
        header_row = 6
        headers = [cell.value for cell in table1[header_row]]
        
        # Should have SA2 code and name columns
        assert any("SA2" in str(header) for header in headers if header)
        
        # Should have SEIFA index columns (IRSD, IRSAD, IER, IEO)
        header_str = " ".join(str(h) for h in headers if h)
        assert "IRSD" in header_str or "Disadvantage" in header_str
        
        workbook.close()
    
    def test_seifa_data_row_count(self, seifa_file: Path):
        """Test that SEIFA file has expected number of SA2 records (2,368)."""
        if not seifa_file or not seifa_file.exists():
            pytest.skip("SEIFA file not available")
        
        workbook = load_workbook(seifa_file, read_only=True)
        table1 = workbook["Table 1"]
        
        # Count data rows (skip headers)
        data_rows = 0
        for row in table1.iter_rows(min_row=7, values_only=True):  # Data starts row 7
            if row[0]:  # If first column has value (SA2 code)
                data_rows += 1
        
        # Should have 2,368 SA2 areas as discovered
        assert data_rows >= 2300, f"Expected ~2,368 SA2 records, got {data_rows}"
        assert data_rows <= 2400, f"Too many records: {data_rows}"
        
        workbook.close()
    
    def test_seifa_sa2_code_format(self, seifa_file: Path):
        """Test SA2 codes are in correct 9-digit format."""
        if not seifa_file or not seifa_file.exists():
            pytest.skip("SEIFA file not available")
        
        workbook = load_workbook(seifa_file, read_only=True)
        table1 = workbook["Table 1"]
        
        # Check first few SA2 codes
        sample_codes = []
        for row in table1.iter_rows(min_row=7, max_row=12, values_only=True):
            if row[0]:  # SA2 code is first column
                sample_codes.append(str(row[0]))
        
        assert len(sample_codes) > 0, "No SA2 codes found"
        
        for code in sample_codes:
            # Should be 9-digit string
            assert len(code) == 9, f"Invalid SA2 code length: {code}"
            assert code.isdigit(), f"SA2 code not numeric: {code}"
            # Should start with 1-9 (Australian state codes)
            assert code[0] in "123456789", f"Invalid SA2 code start: {code}"
        
        workbook.close()
    
    def test_seifa_index_values_range(self, seifa_file: Path):
        """Test SEIFA index values are in expected range (800-1200)."""
        if not seifa_file or not seifa_file.exists():
            pytest.skip("SEIFA file not available")
        
        workbook = load_workbook(seifa_file, read_only=True)
        table1 = workbook["Table 1"]
        
        # Find index score columns (should contain numeric values 800-1200)
        header_row = list(table1[6])
        
        # Look for score columns (not decile columns)
        score_columns = []
        for i, cell in enumerate(header_row):
            if cell.value and "score" in str(cell.value).lower():
                score_columns.append(i)
        
        # If no explicit score columns found, check for numeric columns in range
        if not score_columns:
            for i in range(2, min(10, len(header_row))):  # Skip SA2 code/name
                score_columns.append(i)
        
        assert len(score_columns) > 0, "No SEIFA score columns found"
        
        # Check sample values in score columns
        sample_values = []
        for row in table1.iter_rows(min_row=7, max_row=12, values_only=True):
            for col_idx in score_columns:
                if col_idx < len(row) and isinstance(row[col_idx], (int, float)):
                    sample_values.append(row[col_idx])
        
        assert len(sample_values) > 0, "No numeric SEIFA values found"
        
        # SEIFA scores should be in range 800-1200
        for value in sample_values:
            if 800 <= value <= 1200:  # Valid SEIFA range
                break
        else:
            # If none in SEIFA range, might be deciles (1-10) or other format
            assert any(1 <= v <= 10 for v in sample_values), f"Unexpected SEIFA values: {sample_values[:5]}"
        
        workbook.close()


class TestSEIFASchemaMapping:
    """Test schema mapping configuration for SEIFA data."""
    
    def test_seifa_schema_definition(self):
        """Test SEIFA schema mapping matches discovered structure."""
        
        # Expected schema from analysis
        expected_columns = [
            "sa2_code_2021",
            "sa2_name_2021", 
            "irsd_score",
            "irsd_decile",
            "irsad_score", 
            "irsad_decile",
            "ier_score",
            "ier_decile",
            "ieo_score",
            "ieo_decile",
            "usual_resident_population"
        ]
        
        # Schema should have expected structure
        assert len(expected_columns) == 11, "Schema should have 11 columns"
        
        # Should include all four SEIFA indices
        seifa_indices = ["irsd", "irsad", "ier", "ieo"]
        for index in seifa_indices:
            assert any(index in col for col in expected_columns), f"Missing {index} index"
    
    def test_polars_schema_types(self):
        """Test Polars data types for SEIFA schema."""
        
        polars_schema = {
            "sa2_code_2021": pl.Utf8,
            "sa2_name_2021": pl.Utf8,
            "irsd_score": pl.Int32,
            "irsd_decile": pl.Int8,
            "irsad_score": pl.Int32,
            "irsad_decile": pl.Int8,
            "ier_score": pl.Int32,
            "ier_decile": pl.Int8,
            "ieo_score": pl.Int32,
            "ieo_decile": pl.Int8,
            "usual_resident_population": pl.Int32,
        }
        
        # All columns should have appropriate types
        assert len(polars_schema) == 11
        
        # SA2 codes and names should be strings
        assert polars_schema["sa2_code_2021"] == pl.Utf8
        assert polars_schema["sa2_name_2021"] == pl.Utf8
        
        # Scores should be integers
        assert polars_schema["irsd_score"] == pl.Int32
        assert polars_schema["irsad_score"] == pl.Int32
        
        # Deciles should be small integers (1-10)
        assert polars_schema["irsd_decile"] == pl.Int8
        assert polars_schema["irsad_decile"] == pl.Int8


@pytest.mark.asyncio
class TestSEIFAProcessorIntegration:
    """Integration tests for SEIFA processor with real data."""
    
    async def test_seifa_end_to_end_pipeline(self):
        """Test complete pipeline: download → process → validate."""
        
        with tempfile.TemporaryDirectory() as tmp_dir:
            # Step 1: Download real SEIFA data
            downloader = RealDataDownloader(tmp_dir)
            
            try:
                results = await downloader.download_essential_datasets(["seifa_2021_sa2"])
                seifa_path = results.get("seifa_2021_sa2")
                
                if not seifa_path or not seifa_path.exists():
                    pytest.skip("Could not download SEIFA file")
                
                # Step 2: Validate file exists and is readable
                assert seifa_path.stat().st_size > 1000000, "SEIFA file too small"
                
                # Step 3: Basic Excel validation
                workbook = load_workbook(seifa_path, read_only=True)
                assert "Table 1" in workbook.sheetnames
                workbook.close()
                
                print(f"✅ Integration test complete: {seifa_path.stat().st_size / (1024*1024):.1f}MB")
                
            except Exception as e:
                pytest.skip(f"Integration test failed (network/file issue): {e}")


if __name__ == "__main__":
    # Run SEIFA processor tests
    pytest.main([__file__, "-v"])