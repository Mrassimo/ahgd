"""
Test suite for real Australian data source downloads.

Tests validate that all discovered URLs work correctly and 
files can be downloaded and processed as expected.
"""

import asyncio
import tempfile
import zipfile
from pathlib import Path
from typing import Dict, Optional

import httpx
import pytest
from openpyxl import load_workbook

# Real data source URLs (verified working)
REAL_DATA_SOURCES = {
    # ABS Digital Boundaries
    "sa2_boundaries_gda2020": {
        "url": "https://www.abs.gov.au/statistics/standards/australian-statistical-geography-standard-asgs-edition-3/jul2021-jun2026/access-and-downloads/digital-boundary-files/SA2_2021_AUST_SHP_GDA2020.zip",
        "expected_size_mb": 96,
        "format": "zip",
        "contains": ["SA2_2021_AUST_GDA2020.shp"]
    },
    
    "sa2_boundaries_gda94": {
        "url": "https://www.abs.gov.au/statistics/standards/australian-statistical-geography-standard-asgs-edition-3/jul2021-jun2026/access-and-downloads/digital-boundary-files/SA2_2021_AUST_SHP_GDA94.zip", 
        "expected_size_mb": 47,
        "format": "zip",
        "contains": ["SA2_2021_AUST_GDA94.shp"]
    },
    
    # SEIFA 2021 Data
    "seifa_2021_sa2": {
        "url": "https://www.abs.gov.au/statistics/people/people-and-communities/socio-economic-indexes-areas-seifa-australia/2021/Statistical%20Area%20Level%202%2C%20Indexes%2C%20SEIFA%202021.xlsx",
        "expected_size_mb": 1.3,
        "format": "xlsx",
        "expected_sheets": ["Contents", "Table 1"]  # Real sheet names discovered
    },
    
    # Medicare Benefits Schedule
    "mbs_historical": {
        "url": "https://data.gov.au/data/dataset/8a19a28f-35b0-4035-8cd5-5b611b3cfa6f/resource/519b55ab-8f81-47d1-a483-8495668e38d8/download/mbs-demographics-historical-1993-2015.zip",
        "expected_size_mb": 50,  # Estimate
        "format": "zip",
        "contains": [".csv"]  # Should contain CSV files
    },
    
    # Pharmaceutical Benefits Scheme
    "pbs_current": {
        "url": "https://data.gov.au/data/dataset/14b536d4-eb6a-485d-bf87-2e6e77ddbac1/resource/08eda5ab-01c0-4c94-8b1a-157bcffe80d3/download/pbs-item-2016csvjuly.csv",
        "expected_size_mb": 10,  # Estimate
        "format": "csv"
    },
    
    "pbs_historical": {
        "url": "https://data.gov.au/data/dataset/14b536d4-eb6a-485d-bf87-2e6e77ddbac1/resource/56f87bbb-a7cb-4cbf-a723-7aec22996eee/download/csv-pbs-item-historical-1992-2014.zip",
        "expected_size_mb": 25,  # Estimate
        "format": "zip",
        "contains": [".csv"]
    }
}


class TestRealDataSources:
    """Test suite for real Australian data source validation."""
    
    @pytest.mark.asyncio
    async def test_url_accessibility(self):
        """Test that all real URLs are accessible via HEAD requests."""
        
        async with httpx.AsyncClient(timeout=30.0) as client:
            for source_name, source_info in REAL_DATA_SOURCES.items():
                
                try:
                    # Test with HEAD request (faster, doesn't download)
                    response = await client.head(source_info["url"])
                    
                    # Should not be 404
                    assert response.status_code != 404, f"URL not found: {source_name}"
                    
                    # Should allow GET requests
                    assert response.status_code in [200, 302, 301], f"Unexpected status for {source_name}: {response.status_code}"
                    
                    print(f"✅ {source_name}: Status {response.status_code}")
                    
                except httpx.RequestError as e:
                    pytest.fail(f"Network error accessing {source_name}: {e}")
    
    @pytest.mark.asyncio
    async def test_file_download_small_sample(self):
        """Test downloading first few KB of each file to validate format."""
        
        async with httpx.AsyncClient(timeout=60.0) as client:
            for source_name, source_info in REAL_DATA_SOURCES.items():
                
                try:
                    # Download first 1KB to check file format
                    headers = {"Range": "bytes=0-1023"}
                    response = await client.get(source_info["url"], headers=headers)
                    
                    if response.status_code in [200, 206]:  # 206 = Partial Content
                        content = response.content
                        
                        # Validate file format based on magic bytes
                        if source_info["format"] == "zip":
                            assert content.startswith(b"PK"), f"Not a ZIP file: {source_name}"
                        elif source_info["format"] == "xlsx":
                            assert content.startswith(b"PK"), f"Not an Excel file: {source_name}"
                        elif source_info["format"] == "csv":
                            # CSV should be text-like
                            try:
                                content.decode('utf-8')
                            except UnicodeDecodeError:
                                try:
                                    content.decode('latin-1')
                                except UnicodeDecodeError:
                                    pytest.fail(f"Not a text file: {source_name}")
                        
                        print(f"✅ {source_name}: Format validated ({source_info['format']})")
                    
                    else:
                        print(f"⚠️  {source_name}: Partial download not supported (Status {response.status_code})")
                
                except httpx.RequestError as e:
                    pytest.fail(f"Error testing {source_name}: {e}")
    
    @pytest.mark.asyncio
    @pytest.mark.slow  # Mark as slow test since it downloads full files
    async def test_full_file_download(self):
        """Test full file download for smaller files only."""
        
        # Only test smaller files to avoid long test times
        small_files = {
            k: v for k, v in REAL_DATA_SOURCES.items() 
            if v["expected_size_mb"] < 10
        }
        
        async with httpx.AsyncClient(timeout=120.0) as client:
            for source_name, source_info in small_files.items():
                
                # Use proper file extension for temporary files
                file_extension = f".{source_info['format']}"
                with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
                    try:
                        print(f"Downloading {source_name}...")
                        
                        async with client.stream('GET', source_info["url"]) as response:
                            response.raise_for_status()
                            
                            total_size = 0
                            async for chunk in response.aiter_bytes(chunk_size=8192):
                                tmp_file.write(chunk)
                                total_size += len(chunk)
                        
                        tmp_file.flush()
                        file_path = Path(tmp_file.name)
                        
                        # Validate file size is reasonable
                        actual_size_mb = file_path.stat().st_size / (1024 * 1024)
                        expected_size = source_info["expected_size_mb"]
                        
                        # Allow 50% variance in size estimates
                        assert actual_size_mb > expected_size * 0.5, f"File too small: {source_name}"
                        assert actual_size_mb < expected_size * 2.0, f"File too large: {source_name}"
                        
                        # Validate file contents based on format
                        await self._validate_file_content(file_path, source_info, source_name)
                        
                        print(f"✅ {source_name}: Downloaded and validated ({actual_size_mb:.1f}MB)")
                        
                    except Exception as e:
                        pytest.fail(f"Failed to download {source_name}: {e}")
                    
                    finally:
                        # Cleanup
                        if file_path.exists():
                            file_path.unlink()
    
    async def _validate_file_content(self, file_path: Path, source_info: Dict, source_name: str):
        """Validate file content based on expected format."""
        
        if source_info["format"] == "zip":
            # Test ZIP file can be opened and contains expected files
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                file_list = zip_ref.namelist()
                
                if "contains" in source_info:
                    for expected_file in source_info["contains"]:
                        matching_files = [f for f in file_list if expected_file in f]
                        assert len(matching_files) > 0, f"Expected file type {expected_file} not found in {source_name}"
        
        elif source_info["format"] == "xlsx":
            # Test Excel file can be opened
            try:
                workbook = load_workbook(file_path, read_only=True)
                sheet_names = workbook.sheetnames
                
                if "expected_sheets" in source_info:
                    for expected_sheet in source_info["expected_sheets"]:
                        assert expected_sheet in sheet_names, f"Expected sheet {expected_sheet} not found in {source_name}"
                
                workbook.close()
            except Exception as e:
                pytest.fail(f"Cannot read Excel file {source_name}: {e}")
        
        elif source_info["format"] == "csv":
            # Test CSV file can be read
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    first_line = f.readline()
                    assert len(first_line) > 0, f"Empty CSV file: {source_name}"
                    # Should contain comma separators
                    assert ',' in first_line, f"Not a proper CSV file: {source_name}"
            except UnicodeDecodeError:
                # Try with different encoding
                try:
                    with open(file_path, 'r', encoding='latin-1') as f:
                        first_line = f.readline()
                        assert len(first_line) > 0, f"Empty CSV file: {source_name}"
                except Exception as e:
                    pytest.fail(f"Cannot read CSV file {source_name}: {e}")


class TestDataSourceIntegration:
    """Test integration of data sources with processing pipeline."""
    
    def test_url_format_compatibility(self):
        """Test that URLs are compatible with httpx client."""
        
        for source_name, source_info in REAL_DATA_SOURCES.items():
            url = source_info["url"]
            
            # URL should be properly formatted
            assert url.startswith("https://"), f"Insecure URL: {source_name}"
            
            # URL should not have spaces (except in encoded form)
            if " " in url:
                assert "%20" in url, f"Unencoded spaces in URL: {source_name}"
            
            # URL should be parseable by httpx
            try:
                httpx.URL(url)
            except Exception as e:
                pytest.fail(f"Invalid URL format for {source_name}: {e}")
    
    def test_expected_data_coverage(self):
        """Test that we have all required data source types."""
        
        source_names = set(REAL_DATA_SOURCES.keys())
        
        # Should have geographic boundaries
        geographic_sources = {s for s in source_names if "boundaries" in s}
        assert len(geographic_sources) > 0, "Missing geographic boundary data"
        
        # Should have socio-economic data
        seifa_sources = {s for s in source_names if "seifa" in s}
        assert len(seifa_sources) > 0, "Missing SEIFA socio-economic data"
        
        # Should have health service data
        health_sources = {s for s in source_names if "mbs" in s or "pbs" in s}
        assert len(health_sources) > 0, "Missing health service data"
        
        print(f"✅ Data coverage validated: {len(source_names)} sources")


if __name__ == "__main__":
    # Run basic connectivity test
    asyncio.run(TestRealDataSources().test_url_accessibility())