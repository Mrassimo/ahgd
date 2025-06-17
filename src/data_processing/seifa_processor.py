"""
SEIFA (Socio-Economic Indexes for Areas) processor for real ABS Excel data.

Processes official SEIFA 2021 Excel files with four socio-economic indices:
- IRSD: Index of Relative Socio-economic Disadvantage  
- IRSAD: Index of Advantage and Disadvantage
- IER: Index of Economic Resources
- IEO: Index of Education and Occupation

Based on real data analysis: 2,368 SA2 areas, Excel Table 1 format.
"""

from pathlib import Path
from typing import Dict, List, Optional, Union

import polars as pl
from loguru import logger
from openpyxl import load_workbook
from rich.console import Console

console = Console()

# SEIFA processing configuration based on real data analysis
SEIFA_CONFIG = {
    "filename": "SEIFA_2021_SA2_Indexes.xlsx",
    "primary_sheet": "Table 1",
    "header_row": 6,  # Headers start at row 6 (0-based: row 5)
    "data_start_row": 7,  # Data starts at row 7 (0-based: row 6) 
    "expected_records": 2368,  # Total SA2 areas in Australia
    "expected_sheets": ["Contents", "Table 1", "Table 2", "Table 3", "Table 4", "Table 5", "Table 6"]
}

# Schema mapping based on real Excel structure analysis
SEIFA_SCHEMA = {
    "sa2_code_2021": pl.Utf8,
    "sa2_name_2021": pl.Utf8,
    "irsd_score": pl.Int32,
    "irsd_decile": pl.Int8,
    "irsad_score": pl.Int32,
    "irsad_decile": pl.Int8,
    "ier_score": pl.Int32,
    "ier_decile": pl.Int8,
    "ieo_score": pl.Int32,
    "ieo_decile": pl.Int8,
    "usual_resident_population": pl.Int32,
}

# SEIFA index descriptions for documentation
SEIFA_INDEX_DESCRIPTIONS = {
    "irsd": "Index of Relative Socio-economic Disadvantage - measures disadvantage (low income, unemployment)",
    "irsad": "Index of Advantage and Disadvantage - comprehensive measure of socio-economic status",
    "ier": "Index of Economic Resources - measures household income and dwelling costs",
    "ieo": "Index of Education and Occupation - measures qualifications and skilled employment"
}


class SEIFAProcessor:
    """
    High-performance SEIFA data processor using real Excel schema.
    
    Converts ABS SEIFA 2021 Excel data to analysis-ready Polars DataFrames
    with comprehensive validation and error handling.
    """
    
    def __init__(self, data_dir: Union[str, Path] = "data"):
        self.data_dir = Path(data_dir)
        self.raw_dir = self.data_dir / "raw"
        self.processed_dir = self.data_dir / "processed"
        
        self.raw_dir.mkdir(parents=True, exist_ok=True)
        self.processed_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info("SEIFA Processor initialized")
        logger.info(f"Raw data directory: {self.raw_dir}")
        logger.info(f"Processed data directory: {self.processed_dir}")
    
    def validate_seifa_file(self, file_path: Path) -> bool:
        """
        Validate SEIFA Excel file structure matches expected format.
        
        Args:
            file_path: Path to SEIFA Excel file
            
        Returns:
            True if file is valid, False otherwise
        """
        try:
            if not file_path.exists():
                logger.error(f"SEIFA file not found: {file_path}")
                return False
            
            # Check file size (should be ~1.3MB)
            file_size_mb = file_path.stat().st_size / (1024 * 1024)
            if file_size_mb < 0.5 or file_size_mb > 5.0:
                logger.warning(f"Unexpected SEIFA file size: {file_size_mb:.1f}MB")
            
            # Validate Excel structure
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            
            # Check primary sheet exists
            if SEIFA_CONFIG["primary_sheet"] not in sheet_names:
                logger.error(f"Primary sheet '{SEIFA_CONFIG['primary_sheet']}' not found")
                workbook.close()
                return False
            
            # Check expected sheets
            missing_sheets = []
            for expected_sheet in SEIFA_CONFIG["expected_sheets"]:
                if expected_sheet not in sheet_names:
                    missing_sheets.append(expected_sheet)
            
            if missing_sheets:
                logger.warning(f"Missing expected sheets: {missing_sheets}")
            
            # Validate data structure
            table1 = workbook[SEIFA_CONFIG["primary_sheet"]]
            
            # Check has sufficient rows
            row_count = table1.max_row
            if row_count < SEIFA_CONFIG["expected_records"]:
                logger.warning(f"Fewer rows than expected: {row_count} vs {SEIFA_CONFIG['expected_records']}")
            
            workbook.close()
            logger.info(f"✓ SEIFA file validation passed: {file_path.name}")
            return True
            
        except Exception as e:
            logger.error(f"SEIFA file validation failed: {e}")
            return False
    
    def extract_seifa_data(self, file_path: Path) -> pl.DataFrame:
        """
        Extract SEIFA data from Excel file to Polars DataFrame.
        
        Args:
            file_path: Path to SEIFA Excel file
            
        Returns:
            Polars DataFrame with standardized SEIFA schema
        """
        if not self.validate_seifa_file(file_path):
            raise ValueError(f"Invalid SEIFA file: {file_path}")
        
        logger.info(f"Extracting SEIFA data from {file_path.name}")
        
        try:
            # Read Excel file with openpyxl for better control
            workbook = load_workbook(file_path, read_only=True)
            table1 = workbook[SEIFA_CONFIG["primary_sheet"]]
            
            # Extract headers
            header_row_idx = SEIFA_CONFIG["header_row"]
            headers = []
            for cell in table1[header_row_idx]:
                header_value = cell.value
                if header_value:
                    headers.append(str(header_value).strip())
                else:
                    headers.append("")
            
            # Extract data rows with data cleaning
            data_rows = []
            data_start = SEIFA_CONFIG["data_start_row"]
            
            for row in table1.iter_rows(min_row=data_start, values_only=True):
                if row[0]:  # If first column has value
                    # Check if this is a valid SA2 code (9-digit number)
                    first_col = str(row[0]).strip()
                    if len(first_col) == 9 and first_col.isdigit():
                        # Clean row data - replace "-" and empty values with None
                        cleaned_row = []
                        for cell_value in row:
                            if cell_value == "-" or cell_value == "":
                                cleaned_row.append(None)
                            else:
                                cleaned_row.append(cell_value)
                        data_rows.append(cleaned_row)
                    else:
                        # Stop processing when we hit non-SA2 data (like copyright notices)
                        logger.info(f"Stopping extraction at non-SA2 row: {first_col}")
                        break
            
            workbook.close()
            
            logger.info(f"Extracted {len(data_rows)} data rows with {len(headers)} columns")
            
            # Create unique column names to handle duplicates
            string_headers = []
            seen_headers = {}
            for i, h in enumerate(headers):
                base_name = str(h) if h else f"col_{i}"
                if base_name in seen_headers:
                    seen_headers[base_name] += 1
                    unique_name = f"{base_name}_{seen_headers[base_name]}"
                else:
                    seen_headers[base_name] = 0
                    unique_name = base_name
                string_headers.append(unique_name)
            
            df = pl.DataFrame(data_rows, schema=string_headers, orient="row")
            
            # Standardize column names and types
            standardized_df = self._standardize_seifa_columns(df)
            
            # Validate data quality
            validated_df = self._validate_seifa_data(standardized_df)
            
            logger.info(f"✓ Successfully processed SEIFA data: {len(validated_df)} SA2 areas")
            return validated_df
            
        except Exception as e:
            logger.error(f"Failed to extract SEIFA data: {e}")
            raise
    
    def _standardize_seifa_columns(self, df: pl.DataFrame) -> pl.DataFrame:
        """
        Standardize SEIFA column names and data types.
        
        Maps real Excel column names to our standard schema using position-based mapping
        since the Excel file has a fixed structure.
        """
        logger.info("Standardizing SEIFA column names and types")
        
        # Position-based mapping for SEIFA Excel file structure
        # Based on analysis: columns have fixed positions regardless of header names
        POSITION_MAPPING = {
            0: "sa2_code_2021",              # Column 1: SA2 Code
            1: "sa2_name_2021",              # Column 2: SA2 Name
            2: "irsd_score",                 # Column 3: IRSD Score
            3: "irsd_decile",                # Column 4: IRSD Decile
            4: "irsad_score",                # Column 5: IRSAD Score
            5: "irsad_decile",               # Column 6: IRSAD Decile
            6: "ier_score",                  # Column 7: IER Score
            7: "ier_decile",                 # Column 8: IER Decile
            8: "ieo_score",                  # Column 9: IEO Score
            9: "ieo_decile",                 # Column 10: IEO Decile
            10: "usual_resident_population", # Column 11: Population
        }
        
        # Get the first 11 columns (the ones we need) and create standardized column names
        original_columns = df.columns
        if len(original_columns) < 11:
            raise ValueError(f"Expected at least 11 columns, got {len(original_columns)}")
        
        # Select only the first 11 columns and rename them directly
        new_column_names = []
        for i in range(11):
            if i in POSITION_MAPPING:
                new_column_names.append(POSITION_MAPPING[i])
            else:
                new_column_names.append(f"col_{i}")
        
        # Select first 11 columns and rename them
        standardized_df = df.select(df.columns[:11]).rename(dict(zip(df.columns[:11], new_column_names)))
        
        logger.info(f"Selected and renamed {len(new_column_names)} columns using position-based mapping")
        
        # Apply data type conversions
        type_conversions = {}
        for col in standardized_df.columns:
            if col in SEIFA_SCHEMA:
                target_type = SEIFA_SCHEMA[col]
                type_conversions[col] = target_type
        
        # Convert types with error handling and null handling
        for col, target_type in type_conversions.items():
            try:
                if col in standardized_df.columns:
                    # Handle null values and type conversion
                    if target_type in [pl.Int32, pl.Int8]:
                        # For integer columns, convert nulls and handle string representations
                        standardized_df = standardized_df.with_columns(
                            pl.col(col)
                            .map_elements(lambda x: None if x is None or str(x).strip() in ["-", "", "nan"] else x, return_dtype=pl.Object)
                            .cast(target_type, strict=False)
                            .alias(col)
                        )
                    else:
                        # For string columns
                        standardized_df = standardized_df.with_columns(
                            pl.col(col).cast(target_type, strict=False).alias(col)
                        )
            except Exception as e:
                logger.warning(f"Failed to convert column {col} to {target_type}: {e}")
        
        # Select only standardized columns that exist
        available_columns = [col for col in SEIFA_SCHEMA.keys() if col in standardized_df.columns]
        final_df = standardized_df.select(available_columns)
        
        logger.info(f"Standardized {len(available_columns)} columns: {available_columns}")
        return final_df
    
    def _validate_seifa_data(self, df: pl.DataFrame) -> pl.DataFrame:
        """
        Validate SEIFA data quality and consistency.
        
        Checks SA2 codes, index ranges, and data completeness.
        """
        logger.info("Validating SEIFA data quality")
        
        initial_count = len(df)
        
        # Validate SA2 codes (should be 9-digit strings)
        if "sa2_code_2021" in df.columns:
            df = df.filter(
                pl.col("sa2_code_2021").str.len_chars() == 9
            )
            logger.info(f"Filtered to valid 9-digit SA2 codes: {len(df)}")
        
        # Validate SEIFA scores (should be in range 800-1200, allow nulls)
        score_columns = [col for col in df.columns if "score" in col]
        for score_col in score_columns:
            initial_scores = len(df)
            df = df.filter(
                pl.col(score_col).is_null() | 
                ((pl.col(score_col) >= 800) & (pl.col(score_col) <= 1200))
            )
            if len(df) < initial_scores:
                logger.info(f"Filtered {score_col} to valid range (800-1200): {len(df)}")
        
        # Validate deciles (should be 1-10, allow nulls)
        decile_columns = [col for col in df.columns if "decile" in col]
        for decile_col in decile_columns:
            initial_deciles = len(df)
            df = df.filter(
                pl.col(decile_col).is_null() |
                ((pl.col(decile_col) >= 1) & (pl.col(decile_col) <= 10))
            )
            if len(df) < initial_deciles:
                logger.info(f"Filtered {decile_col} to valid range (1-10): {len(df)}")
        
        # Remove rows with missing critical data
        critical_columns = ["sa2_code_2021"]
        if "sa2_code_2021" in df.columns:
            df = df.filter(pl.col("sa2_code_2021").is_not_null())
        
        final_count = len(df)
        logger.info(f"Data validation complete: {initial_count} → {final_count} records")
        
        if final_count < SEIFA_CONFIG["expected_records"] * 0.9:
            logger.warning(f"Significant data loss during validation: {final_count} vs expected {SEIFA_CONFIG['expected_records']}")
        
        return df
    
    def process_seifa_file(self, filename: Optional[str] = None) -> pl.DataFrame:
        """
        Process SEIFA Excel file to analysis-ready DataFrame.
        
        Args:
            filename: SEIFA Excel filename (default: from config)
            
        Returns:
            Processed Polars DataFrame with SEIFA data
        """
        if filename is None:
            filename = SEIFA_CONFIG["filename"]
        
        file_path = self.raw_dir / filename
        
        if not file_path.exists():
            raise FileNotFoundError(f"SEIFA file not found: {file_path}")
        
        # Extract and process data
        seifa_df = self.extract_seifa_data(file_path)
        
        # Export processed data
        output_path = self.processed_dir / "seifa_2021_sa2.csv"
        seifa_df.write_csv(output_path)
        logger.info(f"Exported processed SEIFA data to {output_path}")
        
        # Also export as Parquet for performance
        parquet_path = self.processed_dir / "seifa_2021_sa2.parquet"
        seifa_df.write_parquet(parquet_path)
        logger.info(f"Exported SEIFA data as Parquet to {parquet_path}")
        
        return seifa_df
    
    def get_seifa_summary(self, seifa_df: pl.DataFrame) -> Dict:
        """
        Generate summary statistics for SEIFA data.
        
        Args:
            seifa_df: Processed SEIFA DataFrame
            
        Returns:
            Dictionary with summary statistics
        """
        summary = {
            "total_sa2_areas": len(seifa_df),
            "states_covered": [],
            "seifa_statistics": {}
        }
        
        # Extract state information from SA2 codes
        if "sa2_code_2021" in seifa_df.columns:
            state_codes = seifa_df.select(
                pl.col("sa2_code_2021").str.slice(0, 1).alias("state_code")
            ).unique()["state_code"].to_list()
            
            state_mapping = {
                "1": "NSW", "2": "VIC", "3": "QLD", "4": "SA", 
                "5": "WA", "6": "TAS", "7": "NT", "8": "ACT"
            }
            
            summary["states_covered"] = [state_mapping.get(code, f"Unknown({code})") for code in state_codes]
        
        # Calculate SEIFA index statistics
        score_columns = [col for col in seifa_df.columns if "score" in col]
        for score_col in score_columns:
            index_name = score_col.replace("_score", "").upper()
            
            stats = seifa_df.select([
                pl.col(score_col).min().alias("min"),
                pl.col(score_col).max().alias("max"), 
                pl.col(score_col).mean().alias("mean"),
                pl.col(score_col).median().alias("median"),
            ]).to_dicts()[0]
            
            summary["seifa_statistics"][index_name] = {
                "min_score": stats["min"],
                "max_score": stats["max"],
                "mean_score": round(stats["mean"], 1) if stats["mean"] else None,
                "median_score": stats["median"],
            }
        
        return summary
    
    def process_complete_pipeline(self) -> pl.DataFrame:
        """
        Execute complete SEIFA processing pipeline.
        
        Returns:
            Fully processed SEIFA DataFrame ready for analysis
        """
        console.print("🏗️  [bold blue]Starting SEIFA processing pipeline...[/bold blue]")
        
        try:
            # Process SEIFA data
            seifa_df = self.process_seifa_file()
            console.print(f"✅ Processed {len(seifa_df)} SA2 areas")
            
            # Generate summary
            summary = self.get_seifa_summary(seifa_df)
            console.print(f"✅ Coverage: {summary['states_covered']}")
            console.print(f"✅ SEIFA indices: {list(summary['seifa_statistics'].keys())}")
            
            console.print("🎉 [bold green]SEIFA processing pipeline complete![/bold green]")
            return seifa_df
            
        except Exception as e:
            console.print(f"❌ [bold red]SEIFA processing failed: {e}[/bold red]")
            raise